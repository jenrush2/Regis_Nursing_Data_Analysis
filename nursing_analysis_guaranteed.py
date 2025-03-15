import pandas as pd
from data_processing import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Are students GUARANTEED to be admitted to the BSN program (no transfer credits for science classes allowed)
# They’ve taken all 4 science classes at Regis
# They’ve completed RCC200
# They have a Science GPA of 3.25 or higher
# They have a cumulative GPA of 3.25 or higher
# They have a C or higher in all classes
# They haven’t withdrawn from any classes

# Load and preprocess data
file_path = "/Users/jenniferrush/Python/Regis_Nursing_Analysis/data/nursing_data.xlsx"
sheet_name = "IDS - Regis College - Pre-Nursi"
raw_data = load_raw_data(file_path, sheet_name)

# Group by student
grouped_data = raw_data.groupby('Student ID#')
result = []

# Define the cohorts you want to INCLUDE
included_cohorts = {'Fall 2023', 'Spring 2023'}

for student_id, group_data in grouped_data:

    entry_cohort = cohort_check(group_data)

    # Skip processing if the student's cohort is NOT in the inclusion list
    if entry_cohort not in included_cohorts:
        continue


    #first_name = keep_column(group_data, 'First Name')
    #last_name = keep_column(group_data, 'Last Name')
    entry_cohort = cohort_check(group_data)
    gpa = keep_column(group_data, 'Cum GPA')
    science_gpa = calculate_science_gpa(group_data)
    any_low_grade = has_low_grade(group_data)
    classes_below_c = get_classes_below_c(group_data)
    minor = keep_column(group_data, 'Added Minor')
    rcc = rcc_check(group_data)
    science_grade_check_var = science_grade_check(group_data)
    science_below_c = science_below_c_list(group_data)
    science_6_at_regis = science_6_at_regis_check(group_data)
    science_8_at_regis = science_8_at_regis_check(group_data)
    science_remaining_regis = science_at_regis_remaining(group_data)
    science_non_regis = list_of_science_transfer_classes(group_data)
    withdrawn_classes = list_of_withdrawn_classes(group_data)
    registered_remaining = registered_for_remaining_check(group_data)
    guaranteed_check = guaranteed_admission_check(group_data)


    result.append({
        'Student ID#': student_id,
        #'First Name': first_name,
        #'Last Name': last_name,
        'Entry Cohort': entry_cohort,
        'Cumulative GPA': gpa,
        'Science GPA': science_gpa,
        'Guaranteed Admission': guaranteed_check,
        'RCC': rcc,
        'Any Grade Lower Than C': any_low_grade,
        'Science Grade Lower Than C': science_grade_check_var,
        'Completed at least 6 of Req Science at Regis': science_6_at_regis,
        'Completed all 8 of Req Science at Regis': science_8_at_regis,
        'Science Classes Lower Than C': science_below_c,
        'Classes Lower Than C': classes_below_c,
        'Science Classes Remaining (Trans Exc)': science_remaining_regis,
        'List of Science Classes Transferred In': science_non_regis,
        'List of Classes Withdrawn': withdrawn_classes,
        'Registered for Remaining Sci': registered_remaining,
        'Minor': minor
    })

# Convert to DataFrame and save to Excel
nursing_final = pd.DataFrame(result)
nursing_final.to_excel('nursing_report_guaranteeed_adm.xlsx', index=False, engine='openpyxl')


#load the workbook and select active sheet
wb = load_workbook('nursing_report_guaranteeed_adm.xlsx')
ws = wb.active

#define a yellow fill pattern
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

#map column names to their index
header = {cell.value: cell.column for cell in ws[1]}

#loop through columns we want highlighting in
for row in range(2, ws.max_row + 1):  # start from row 2 to skip header
    #adjust column index
    any_grade_cell = ws.cell(row=row, column=header.get('Any Grade Lower Than C')) 
    overall_gpa_cell = ws.cell(row=row, column=header.get("Cumulative GPA"))
    science_grade_cell = ws.cell(row=row, column=header.get('Science Grade Lower Than C'))
    science_gpa_cell = ws.cell(row=row, column=header.get('Science GPA'))
    entry_cohort_cell = ws.cell(row=row, column=header.get('Entry Cohort'))
    guaranteed_admission_cell = ws.cell(row=row, column=header.get('Guaranteed Admission'))

    #apply yellow fill if conditions are met
    if any_grade_cell.value == 'yes':
        any_grade_cell.fill = yellow_fill
    if float(overall_gpa_cell.value) < 3.25:
        overall_gpa_cell.fill = yellow_fill
    if science_grade_cell.value == 'yes':
        science_grade_cell.fill = yellow_fill
    if science_gpa_cell.value < 3.25:
        science_gpa_cell.fill = yellow_fill
    if entry_cohort_cell.value == 'TRANSFER':
        entry_cohort_cell.fill = yellow_fill
    if guaranteed_admission_cell.value == 'no':
        guaranteed_admission_cell.fill = yellow_fill



# save the formatted workbook
wb.save('nursing_final_analysis_highlighted.xlsx')