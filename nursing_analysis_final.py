import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


raw_data = pd.read_excel("/Users/jenniferrush/Python/Regis_Nursing_Analysis/data/nursing_data.xlsx", sheet_name="IDS - Regis College - Pre-Nursi")

#Change all of the data types so they are strings or floats instead of objects
raw_data[['Added Minor', 'Class Level', 'Entry Cohort', 'Enrolled Semester','Dept', 'Course Number', 'Course Title']] = raw_data[['Added Minor', 'Class Level', 'Entry Cohort', 'Enrolled Semester','Dept', 'Course Number', 'Course Title']].astype("string")
raw_data['Verified Grade'] = raw_data['Verified Grade'].replace({'A': '4.000', 'A-': '3.667' ,'B+': '3.333', 'B': '3.000', 'B-': '2.667', 'C+': '2.333', 'C': '2.000', 'C-': '1.667', 'D+': '1.333', 'D': '1.000', 'D-': '0.667', 'F': '0.000', 'W': '7'})
raw_data['Verified Grade'] = pd.to_numeric(raw_data['Verified Grade'], errors = 'coerce', downcast = 'float')
raw_data['Completed Credits'] = pd.to_numeric(raw_data['Completed Credits'], errors = 'coerce', downcast = 'float')


#If a student earned an F, their credits are listed as 0. 
#Take science classes we care about and changes credits to 3 for course and 1 for lab
raw_data.loc[
    (
        (raw_data['Dept'].str.contains('CH')) & 
        (raw_data['Course Number'].str.contains('206A'))
    ) | 
    (
        (raw_data['Dept'].str.contains('BL')) & 
        (raw_data['Course Number'].str.contains('254|274|276')) 
    ) &
    (
        (raw_data['Verified Grade'] == '0.00')
    ), 
    'Completed Credits'
] = 3.000

raw_data.loc[
    (
        (raw_data['Dept'].str.contains('CH')) & 
        (raw_data['Course Number'].str.contains('207A'))
    ) | 
    (
        (raw_data['Dept'].str.contains('BL')) & 
        (raw_data['Course Number'].str.contains('255|275|277')) 
    ) &
    (
        (raw_data['Verified Grade'] == '0.00')
    ),  
    'Completed Credits'
] = 1.000

#Group by Student ID#
grouped_data = raw_data.groupby('Student ID#')

result = []

#Iterate through each group
for student_id, group_data in grouped_data:
    #calculations to create each variable

    #gpa keep the same
    gpa = group_data['Cum GPA'].iloc[0]
    
    #entry cohort
    def cohort_check(entry_data):

        def year(data):
            two_digit_year = data[0:2]
            year = '20' + two_digit_year
            return year

        if 'STR' in entry_data or 'FTR' in entry_data:
            return 'TRANSFER'
        elif 'FN' in entry_data:
            year = year(entry_data)
            return 'Fall ' + year
        elif 'SN' in entry_data:
            year = year(entry_data)
            return 'Spring ' + year
        else:
            return entry_data

    entry_cohort = cohort_check(group_data['Entry Cohort'].iloc[0])

    #science gpa, exclude NaN
    science_courses = group_data.loc[
    (group_data['Dept'].str.contains('BL|CH', na=False)) & 
    group_data['Verified Grade'].notna()
    ]

    weighted_sum = (science_courses['Verified Grade'] * science_courses['Completed Credits']).sum()
    total_credits = science_courses['Completed Credits'].sum()

    #for some reason this was rounding at the beginning, but is no longer rounding...
    science_gpa = (weighted_sum/total_credits).round(2) if total_credits != 0 else 0.00

    

    #earn a c or higher in RCC200
    rcc_check = 'yes' if (((group_data['Dept'].str.contains('RCC')) & (group_data['Course Number'].str.contains('200'))) 
                            & (group_data['Verified Grade'] >= 2)).any() else 'no'


    #any grades lower than a c
    all_grade_check ='yes' if (group_data['Verified Grade'] < 2).any() else 'no'

    #list of classes lower than a c
    classes_below_c = group_data.loc[(group_data['Verified Grade'] < 2) & (group_data['Completed Credits'] > 0.00), ['Dept', 'Course Number']]

    #concat Dept and Course Number
    classes_below_c = classes_below_c['Dept'] + classes_below_c['Course Number']

    #convert list into comma separated string or none
    classes_below_c = ', '.join(classes_below_c.tolist()) if not classes_below_c.empty else ''

    #science grades lower than a c
    science_grade_check = 'yes' if ((group_data['Verified Grade'] < 2) 
                                    & ((group_data['Dept'].str.contains('BL')) | (group_data['Dept'].str.contains('CH')))).any() else 'no'
   
    #list of science classes lower than a c
    science_below_c = group_data.loc[((group_data['Verified Grade']) < 2) & (group_data['Dept'].str.contains('BL') | group_data['Dept'].str.contains('CH')) & (group_data['Completed Credits'] > 0.00), ['Dept', 'Course Number']]
    science_below_c = science_below_c['Dept'] + science_below_c['Course Number']
    science_below_c = ', '.join(science_below_c.tolist()) if not science_below_c.empty else ''
    
    #minor
    minor = group_data['Added Minor'].iloc[0]
    
    # completed 6 out of 8 science courses?
    # Select specific science courses where the student earned C or above
    science_c_or_above = group_data.loc[
    (group_data['Verified Grade'] >= 2) 
    & (
        ((group_data['Dept'].str.contains('CH', na=False)) & 
         (group_data['Course Number'].str.contains('206A|207A', na=False))) 
        | 
        ((group_data['Dept'].str.contains('BL', na=False)) & 
         (group_data['Course Number'].str.contains('254|255|274|275|276|277', na=False)))
    ) 
    & (group_data['Completed Credits'] > 0.00), 
    ['Dept', 'Course Number']
]

    # Concatenate 'Dept' and 'Course Number' columns
    science_c_or_above = science_c_or_above['Dept'] + science_c_or_above['Course Number']

    # Convert list to a string with courses separated by ", " or empty string if no courses
    science_c_or_above = ', '.join(science_c_or_above.tolist()) if not science_c_or_above.empty else ''

    # Count the number of courses completed
    science_6_check = 'yes' if len(science_c_or_above.split(', ')) >= 6 else 'no'

    #completed all 8 science courses? missing which?
    science_8_check = 'yes' if len(science_c_or_above.split(', ')) >= 8 else 'no'

    # List of required science courses
    all_science_classes_list = ['CH206A', 'CH207A', 'BL254', 'BL255', 'BL274', 'BL275', 'BL276', 'BL277']

    # Convert completed courses into a set
    completed_science_courses = set(science_c_or_above.split(', ')) if science_c_or_above else set()

    # Find missing courses
    science_remaining = list(set(all_science_classes_list) - completed_science_courses)
    science_remaining = ', '.join(science_remaining) if science_remaining else ''
    
    #withdrawn from any classes? which?
    withdrawn = group_data.loc[group_data['Verified Grade'] == 7, ['Dept', 'Course Number']]
    withdrawn = withdrawn['Dept'] + withdrawn['Course Number']
    withdrawn = ', '.join(withdrawn.tolist()) if not withdrawn.empty else ''
    
    #registered for remaining science courses?
    #check for courses from the science_remaining list AND empty verfied grade column
    registered_science_classes = group_data.loc[
        (group_data['Verified Grade'].isnull())
        & (
        ((group_data['Dept'].str.contains('CH', na=False)) & 
         (group_data['Course Number'].str.contains('206A|207A', na=False))) 
        | 
        ((group_data['Dept'].str.contains('BL', na=False)) & 
         (group_data['Course Number'].str.contains('254|255|274|275|276|277', na=False)))
        )
        ]
    
    registered_science_classes = registered_science_classes['Dept'] + registered_science_classes['Course Number']
    registered_science_classes = ', '.join(registered_science_classes.tolist()) if not registered_science_classes.empty else ''
    #make sure both string lists are converted to sets
    registered_science_set = set(registered_science_classes.split(', ')) if registered_science_classes else set()
    science_remaining_set = set(science_remaining.split(', ')) if science_remaining else set()
    
    if not science_remaining:
        registered = ''
    else: 
        registered = 'yes' if set(registered_science_classes) == set(science_remaining) else 'no'
        

    #Check with Lynetta for all requirements for admission check(like overal gpa, registered for remaining, etc.)
    #admission check
    admission_check = 'no' if ((entry_cohort == 'TRANSFER') | (science_6_check == 'no') | (science_gpa < 3.25)| (gpa < 3.25) | (rcc_check == 'no')) else ''


    #result as key and value pairs
    result.append(
        {'Student ID#': student_id, 
         'Entry Cohort': entry_cohort,
         'Cumulative GPA': gpa, 
         'Science GPA': science_gpa, 
         'Guaranteed Admission': admission_check, 
         'RCC 200': rcc_check, 
         'Any Grade Lower Than C': all_grade_check, 
         'Science Grade Lower Than C': science_grade_check,
         'Withdrawn': withdrawn,
         '6 of 8 Science Classes Completed': science_6_check,  
         'All 8 Science Classes Completed': science_8_check, 
         'Science Classes Remaining': science_remaining,
         'Registered for Remaining': registered,
         'Science Classes Lower Than C': science_below_c,
         'Classes Lower Than C': classes_below_c,
         'Minor': minor,}
         )

#result to a new dataframe
nursing_final = pd.DataFrame(result)

#drop ID numbers and use openpyxl to export
nursing_final.to_excel('nursing_analysis_final.xlsx', index=False, engine='openpyxl')

#load the workbook and select active sheet
wb = load_workbook('nursing_analysis_final.xlsx')
ws = wb.active

#define a yellow fill pattern
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

#map column names to their index
header = {cell.value: cell.column for cell in ws[1]}

#loop through columns we want highlighting in
for row in range(2, ws.max_row + 1):  # start from row 2 to skip header
    #adjust column index
    any_grade_cell = ws.cell(row=row, column=header.get('Any Grade Lower Than C')) 
    overall_gpa_cell = ws.cell(row=row, column=header.get("Cumulative GPA"))
    science_grade_cell = ws.cell(row=row, column=header.get('Science Grade Lower Than C'))
    science_gpa_cell = ws.cell(row=row, column=header.get('Science GPA'))
    entry_cohort_cell = ws.cell(row=row, column=header.get('Entry Cohort'))
    guaranteed_admission_cell = ws.cell(row=row, column=header.get('Guaranteed Admission'))

    #apply yellow fill if conditions are met
    if any_grade_cell.value == 'yes':
        any_grade_cell.fill = yellow_fill
    if overall_gpa_cell.value < 3.25:
        overall_gpa_cell.fill = yellow_fill
    if science_grade_cell.value == 'yes':
        science_grade_cell.fill = yellow_fill
    if science_gpa_cell.value < 3.25:
        science_gpa_cell.fill = yellow_fill
    if entry_cohort_cell.value == 'TRANSFER':
        entry_cohort_cell.fill = yellow_fill
    if guaranteed_admission_cell.value == 'no':
        guaranteed_admission_cell.fill = yellow_fill



# save the formatted workbook
wb.save('nursing_final_analysis_highlighted.xlsx')

